#!/usr/bin/env python3
"""
Export customer request/order/note data (excluding payment details) to an XLSX file.

Usage:
    python3 scripts/export_customers.py [--output exports/customer_data.xlsx]
"""

from __future__ import annotations

import argparse
import datetime as dt
import os
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import mysql.connector
from mysql.connector.connection import MySQLConnection
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


DEFAULT_SOCKET = "/var/run/mysqld/mysqld.sock"
DEFAULT_DB = "preprod"
DEFAULT_USER = "preprod_user"
DEFAULT_PASSWORD = "!1q2w3eZ"
DEFAULT_OUTPUT = Path("exports/customer_data.xlsx")


def get_connection() -> MySQLConnection:
    return mysql.connector.connect(
        user=DEFAULT_USER,
        password=DEFAULT_PASSWORD,
        unix_socket=DEFAULT_SOCKET,
        database=DEFAULT_DB,
        charset="utf8",
    )


def fetch_customer_rows(conn: MySQLConnection) -> Iterable[Tuple[Any, ...]]:
    query = """
        SELECT
            r.id AS request_id,
            r.request_date,
            r.processed_date,
            r.industry,
            r.industry_send,
            r.ref_source,
            r.other_source,
            r.search_id,
            r.search_keyword,
            r.user_ip,
            r.conversations,
            r.complete_address,
            r.tracking_number,
            r.order_data,
            u.id AS user_id,
            uc.company AS account_company_name,
            uc.abbr AS account_company_abbr,
            u.first_name,
            u.last_name,
            u.email,
            u.email_alt,
            u.phone,
            u.phone_ext,
            u.phone_type,
            u.position,
            u.industry AS customer_industry,
            u.street,
            u.street2,
            u.city,
            u.state,
            u.zipcode,
            u.country,
            u.user_abbr,
            u.admin_comment,
            uj.job_id AS job_code,
            uj.estimate_id,
            uj.order_total,
            uj.order_counts,
            rn.id AS note_id,
            rn.type AS note_type,
            rn.date AS note_date,
            rn.text AS note_text,
            CONCAT(au.first_name, ' ', au.last_name) AS note_author
        FROM requests r
        LEFT JOIN users u ON r.user_id = u.id
        LEFT JOIN users_company uc ON u.company_id = uc.id
        LEFT JOIN user_jobs uj ON r.job_id = uj.id
        LEFT JOIN request_notes rn
            ON rn.request_id = r.id
           AND rn.removed = 0
           AND rn.type NOT IN ('payment', 'credit')
        LEFT JOIN users au ON rn.author_id = au.id
        ORDER BY r.id, rn.date;
    """
    cursor = conn.cursor()
    cursor.execute(query)
    for row in cursor:
        yield row
    cursor.close()


def prepare_workbook(headers: List[str]) -> Workbook:
    wb = Workbook(write_only=False)
    ws = wb.active
    ws.title = "Customers"
    ws.append(headers)
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = min(max(len(header) + 2, 12), 60)
    return wb


def sanitize_string(value: str) -> str:
    return ILLEGAL_CHARACTERS_RE.sub("", value)


def serialize(value: Any) -> Any:
    if isinstance(value, (dt.datetime, dt.date)):
        return value
    if value is None:
        return ""
    if isinstance(value, str):
        return sanitize_string(value)
    return value


def export_to_xlsx(rows: Iterable[Tuple[Any, ...]], output_path: Path) -> int:
    headers = [
        "Request ID",
        "Request Date",
        "Processed Date",
        "Industry",
        "Industry (Sent)",
        "Referral Source",
        "Other Source",
        "Search ID",
        "Search Keyword",
        "User IP",
        "Conversations",
        "Complete Address",
        "Tracking Numbers",
        "Order Data (raw)",
        "Customer ID",
        "Account Company Name",
        "Account Company Abbr",
        "First Name",
        "Last Name",
        "Email",
        "Alt Email",
        "Phone",
        "Phone Ext",
        "Phone Type",
        "Position",
        "Customer Industry",
        "Street",
        "Street 2",
        "City",
        "State",
        "Postal Code",
        "Country",
        "User Abbr",
        "Admin Comment (data-entry flags)",
        "Job Code",
        "Estimate ID",
        "Order Total",
        "Order Count",
        "Note ID",
        "Note Type",
        "Note Date",
        "Note Text",
        "Note Author",
    ]

    wb = prepare_workbook(headers)
    ws = wb.active

    row_count = 0
    for row in rows:
        ws.append([serialize(value) for value in row])
        row_count += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return row_count


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export customer data to XLSX.")
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Path to output XLSX file (default: {DEFAULT_OUTPUT})",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    conn = get_connection()
    try:
        rows = fetch_customer_rows(conn)
        count = export_to_xlsx(rows, args.output)
    finally:
        conn.close()
    print(f"Wrote {count} rows to {args.output.resolve()}")


if __name__ == "__main__":
    main()

